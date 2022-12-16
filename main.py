from tkinter import *
from tkinter import scrolledtext, ttk, messagebox, filedialog
from Symulacja import Symulacja
from SetParameters import SetParameters, SetParametersD, SetParametersM, SetParametersC, SetParametersW, SetParametersBA
import Help
import constants
import numpy as np
import openpyxl
import matplotlib.pyplot as plt


""" stałe globalne"""
zdrowa = constants.ZDROWA_L[:]  # parametry do funkcji prawdopodoństw
uszkodzona = constants.USZKODZONA_L[:]
zmutowana = constants.ZMUTOWANA_L[:]
nowotworowa = constants.NOWOTWOROWA_L[:]
wspolne = constants.WSPOLNE_L[:]
adapt = constants.ADAPT_L[:]
sasiedztwo = constants.SASIEDZTWO_L[:]

""" zmienne globalne"""
wybor = 0  # limfocyty
wczytywanie = 0  # 0 - nie wczytano pliku, 1 - wczytano plik
parametry = [2, 5, 10, 1.0, 0, 4, 0]  # ilość komórek, kroków, skalowanie czasu, nr. sym, promień, dawka stała
dawka = []  # wczytana dawka
organizm = []  # wczytane komórki
zapis = []
org = []

zd, uszk, zm, nw, ws = [], [], [], [], []

wykr = []

# do sprawdzenia, czy dana funkcja jest uwzględniana
fun_z = [1, 1, 1, 1, 1, 0]  # spont śmierć, rozmnażanie, spont uszk, ...
fun_u = [1, 1, 1, 1, 1, 1, 1, 0, 0]
fun_m = [1, 1, 1, 1, 1, 1, 1, 0, 0, 1]
fun_n = [1, 1, 1, 1, 1]

tekst = " "

# NIE DO KOŃCA AKTUALNE PO ZMIANIE INICJOWANIA KOMÓREK
def pomoc(windo):  # okno pomocy
    wind = Toplevel(windo)
    wind.resizable(0, 0)
    wind.title("Pomoc")
    info = Help.pomoc  # wczytanie tekstu z dodatkowego pliku
    scroll = scrolledtext.ScrolledText(wind, padx=10, pady=10)
    scroll.pack(fill=BOTH, expand=1)
    scroll.insert(END, info)
    scroll.config(state=DISABLED)

# DLA UŻYTKOWNIKA
def set_parameters(wind):  # tworzenie okna wyboru parametrów
    global zdrowa, uszkodzona, zmutowana, nowotworowa, wspolne, adapt, sasiedztwo, wybor
    global fun_z, fun_u, fun_m, fun_n

    wybor = list(listBox.curselection())[0]
    windo = Toplevel(wind)  # wywołanie okna wyboru parametrów
    windo.geometry("450x300")
    windo.title("Wybór parametrów")

    bottom_frame2 = Frame(windo)  # stworzenie dolnej ramki z przyciskami
    bottom_frame2.pack(side=BOTTOM, fill=Y)
    bottom_frame2.config(padx=10, pady=5)

    but_save = Button(bottom_frame2, text="Zapisz", command=lambda: get_par(sp1, sp2, sp3, sp4, sp5, sp6))
    but_save.grid(row=0, column=1, padx=10, pady=0)
    but_save.config(width=10)

    but_reset = Button(bottom_frame2, text="Reset", command=lambda: reset(sp1, sp2, sp3, sp4, sp5, sp6))
    but_reset.grid(row=0, column=2, padx=10, pady=0)
    but_reset.config(width=10)

    tab_parent = ttk.Notebook(windo)  # zakładki dla różnych statusów komórek
    tab1 = ttk.Frame(tab_parent)
    tab_parent.add(tab1, text="zdrowa")
    sp1 = SetParameters(tab1, zdrowa, fun_z)
    sp1.on_select()

    tab2 = ttk.Frame(tab_parent)
    tab_parent.add(tab2, text="uszkodzona")
    sp2 = SetParametersD(tab2, uszkodzona, fun_u)
    sp2.on_select()

    tab3 = ttk.Frame(tab_parent)
    tab_parent.add(tab3, text="zmutowana")
    sp3 = SetParametersM(tab3, zmutowana, fun_m)
    sp3.on_select()

    tab4 = ttk.Frame(tab_parent)
    tab_parent.add(tab4, text="nowotworowa")
    sp4 = SetParametersC(tab4, nowotworowa, fun_n)
    sp4.on_select()

    tab5 = ttk.Frame(tab_parent)
    tab_parent.add(tab5, text="wspólne")
    sp5 = SetParametersW(tab5, wspolne)

    tab6 = ttk.Frame(tab_parent)
    tab_parent.add(tab6, text="pozostałe")
    sp6 = SetParametersBA(tab6, sasiedztwo, adapt)

    tab_parent.pack(fill='both')
    windo.mainloop()


def wczytaj_kom():  # wczytywanie tablicy komórek z wcześniejszej symulacji
    global zdrowa, uszkodzona, zmutowana, nowotworowa, wspolne, wczytywanie, organizm  # wczytywane wartości
    file = filedialog.askopenfilename(initialdir=".", filetypes=[("Plik npz", "*.npz")])  # wybór pliku
    filename = str(file)
    if filename != '':
        dane = np.load(filename, allow_pickle=True)  # wczytanie pliku
        try:  # sprawdzenie poprawności zawartości
            zdrowa = dane['z'].tolist()
            uszkodzona = dane['u'].tolist()
            zmutowana = dane['zm'].tolist()
            nowotworowa = dane['n'].tolist()
            wspolne = dane['w'].tolist()
            organizm = dane['org'].tolist()
        except KeyError:
            messagebox.showinfo("uwaga", "Błędna zawartość pliku")
        else:
            pocz_wart.delete(0, END)  # ustawienie proponowanych liczb komórek do symulacji
            pocz_wart.insert(0, len(organizm[0]) - 2)
            wszystkie_wart.delete(0, END)
            wszystkie_wart.insert(0, len(organizm[0]) - 2)
            wczytywanie = 1

            if zdrowa == constants.ZDROWA_L and uszkodzona == constants.USZKODZONA_L and zmutowana == constants.ZMUTOWANA_L and wspolne == constants.WSPOLNE_L and nowotworowa == constants.NOWOTWOROWA_L:
                selection = listBox.curselection()
                for i in selection:
                    listBox.select_clear(i)
                listBox.select_set(0)
            else:
                selection = listBox.curselection()
                for i in selection:
                    listBox.select_clear(i)
                listBox.select_set(1)


def wczytaj_daw():  # wczytywanie dawki z pliku zewnętrznego txt lub xls (w kolejnych wierszach kolejne dawki)
    global dawka  # tablica z dawką
    dawka = []
    file = filedialog.askopenfilename(initialdir=".", filetypes=[("Plik", ["*.txt", "*.xlsx"])])
    filename = str(file)  # wybór pliku

    if filename != '':
        x = filename.rfind('.')
        ext = filename[(x + 1):len(filename)]  # sprawdzenie rozszerzenia
        if ext == 'txt':
            dawka = read_txt(filename)
        if ext == 'xlsx':
            dawka = read_xlsx(filename)
        if dawka:
            kroki.delete(0, END)
            kroki.insert(0, len(dawka)+1)
            dawk.delete(0, END)
            dawk.insert(0, 0)


def read_txt(filename):  # wczytywanie pliku tekstowego
    with open(filename, 'r') as data_file:
        global dawka
        dawka = []
        data = data_file.read().splitlines()  # odczytanie kolejnych linii
        data_file.close()
        for i in range(len(data)):
            try:  # sprawdzenie, czy typ wczytanej wartości jest liczbowy
                dawka.append(float(data[i]))
            except ValueError:
                messagebox.showinfo("uwaga", "Błędna zawartość pliku - upewnij się że nie ma w nim niewłaściwych znaków lub pustch wierszy")
                return []
        if not dawka:
            messagebox.showinfo("uwaga", "Pusty plik")
        return dawka


def read_xlsx(filename):
    arkusz = openpyxl.load_workbook(filename)  # wczytanie pliku
    sheet = arkusz.active  # wczytanie aktywnej karty
    maxx = sheet.max_row
    for row in sheet.iter_rows(max_row=maxx):
        for cell in row:
            if cell.value is None:  # pusta komórka
                messagebox.showinfo("uwaga", "Wczytano " + str(cell.row-1)+" elementów. Plik może nie być kompletny.")
                return dawka
            else:
                try:
                    dawka.append(float(cell.value))
                except ValueError or TypeError:
                    messagebox.showinfo("uwaga", "Błędna zawartość pliku")
                    return []
    return dawka


def get_par(sp1, sp2, sp3, sp4, sp5, sp6):  # zapisanie informacji o wyborze funkcji
    global zdrowa, uszkodzona, zmutowana, nowotworowa, wspolne, adapt, sasiedztwo
    global fun_z, fun_u, fun_m, fun_n

    zdrowa, fun_z = sp1.get_data()
    uszkodzona, fun_u = sp2.get_data()
    zmutowana, fun_m = sp3.get_data()
    nowotworowa, fun_n = sp4.get_data()
    wspolne = sp5.get_data()
    sasiedztwo, adapt = sp6.get_data()


def reset(sp1, sp2, sp3, sp4, sp5, sp6):  # resetuje wartości w klasie setParameters
    global zdrowa, uszkodzona, zmutowana, nowotworowa, wspolne, adapt, sasiedztwo, wybor

    if wybor == 0:
        zdrowa[:] = constants.ZDROWA_L[:]
        uszkodzona[:] = constants.USZKODZONA_L[:]
        zmutowana[:] = constants.ZMUTOWANA_L[:]
        nowotworowa[:] = constants.NOWOTWOROWA_L[:]
        wspolne[:] = constants.WSPOLNE_L[:]
        sasiedztwo[:] = constants.SASIEDZTWO_L[:]
        adapt[:] = constants.ADAPT_L[:]

    sp1.reset(zdrowa)
    sp2.reset(uszkodzona)
    sp3.reset(zmutowana)
    sp4.reset(nowotworowa)
    sp5.reset(wspolne)
    sp6.reset(sasiedztwo, adapt)


def reset2():  # resetuje parametry w oknie głównym
    global parametry, wczytywanie, dawka, organizm
    pocz_wart.delete(0, END)
    wszystkie_wart.delete(0, END)
    kroki.delete(0, END)
    dawk.delete(0, END)
    par.delete(0, END)
    pocz_wart.insert(0, 5)
    wszystkie_wart.insert(0, 7)
    kroki.insert(0, 500)
    dawk.insert(0, 0)
    par.insert(0, 4)
    czas.current(2)
    dawka = []
    organizm = []
    wczytywanie = 0


# pomocnicza, nieużywana
def zapisz_dane2(filename):

    global tekst
    do_zapisu = []

    t = "zdrowe\t"
    do_zapisu.append(zapis[0])
    t = t + "uszkodzone\t"
    do_zapisu.append(zapis[1])
    t = t + "zmutowane\t"
    do_zapisu.append(zapis[2])
    t = t + "nowotworowe\t"
    do_zapisu.append(zapis[3])
    t = t + "martwe\t"
    do_zapisu.append(zapis[4])
    t = t + "żywe\t"
    do_zapisu.append(zapis[5])
    t = t + "napr_spont\t"
    do_zapisu.append(zapis[6])
    t = t + "adapt\t"
    do_zapisu.append(zapis[7])
    t = t + "\n"
    for x in range(len(zapis[0])):
        for y in range(len(do_zapisu)):
            t = t + str(do_zapisu[y][x]) + "\t"
        t = t + "\n"
    with open(filename, 'w') as file:
        file.write(tekst)
        file.write(t)


def zapisz_dane():
    global zapis

    if len(zapis) == 0:
        messagebox.showinfo("uwaga", "Brak danych do wyświetlenia - przeprowadź symulację.")
    else:
        do_zapisu = []
        t = ""
        if var_healthy.get():
            t = "zdrowe\t"
            do_zapisu.append(zapis[0])
        if var_damaged.get():
            t = t + "uszkodzone\t"
            do_zapisu.append(zapis[1])
        if var_mutated.get():
            t = t + "zmutowane\t"
            do_zapisu.append(zapis[2])
        if var_cancerous.get():
            t = t + "nowotworowe\t"
            do_zapisu.append(zapis[3])
        if var_dead.get():
            t = t + "martwe\t"
            do_zapisu.append(zapis[4])
        if var_all.get():
            t = t + "żywe\t"
            do_zapisu.append(zapis[5])
        t = t + "\n"
        for x in range(len(zapis[0])):
            for y in range(len(do_zapisu)):
                t = t + str(do_zapisu[y][x])+"\t"
            t = t + "\n"

        file = filedialog.asksaveasfilename(initialdir=".", filetypes=[("Plik .txt", ["*.txt"])])
        filename = str(file)
        if filename != '':
            x = filename.rfind('.')
            if x == -1:
                filename = filename + ".txt"

            with open(filename, 'w') as file:
                file.write(t)


def zapisz_komorki():
    global org, zd, uszk, zm, nw, ws
    if len(org) == 0:
        messagebox.showinfo("uwaga", "Brak danych do zapisania - przeprowadź symulację.")
    else:
        file = filedialog.asksaveasfilename(initialdir=".", filetypes=[("Plik .npz", ["*.npz"])])
        filename = str(file)
        if filename != '':
            x = filename.rfind('.')
            if x == -1:
                filename = filename + ".npz"

            with open(filename, 'wb') as file:
                np.savez(file, org=org, z=zd, u=uszk, zm=zm, n=nw, w=ws)


def wyswietl_wykres():  # , wykr
    global zapis, parametry, wykr

    plt.close("symulacja nr " + str(parametry[4]))

    if len(zapis) == 0:
        messagebox.showinfo("uwaga", "Brak danych do wyświetlenia - przeprowadź symulację.")
    else:
        plt.figure("symulacja nr " + str(parametry[4]))
        if var_healthy.get():
            plt.plot(zapis[0], label="zdrowe")
        if var_damaged.get():
            plt.plot(zapis[1], label="uszkodzone")
        if var_mutated.get():
            plt.plot(zapis[2], label="zmutowane")
        if var_cancerous.get():
            plt.plot(zapis[3], label="nowotworowe")
        if var_dead.get():
            plt.plot(zapis[4], label="martwe")
        if var_all.get():
            plt.plot(zapis[5], label="żywe")
        plt.legend()
        plt.show()


def dalej():  # wywłuje symulację
    global zdrowa, uszkodzona, zmutowana, nowotworowa, wspolne, adapt, sasiedztwo, wczytywanie, dawka, organizm
    global fun_z, fun_u, fun_m, fun_n, parametry, s, par, zapis, org, zd, uszk, zm, nw, ws, wykr, tekst

    try:  # sprawdzenie poprawności podanych parametrów
        parametry[0] = float(pocz_wart.get())
        parametry[1] = int(wszystkie_wart.get())
        parametry[2] = int(kroki.get())
        parametry[5] = int(par.get())  # promień
        parametry[6] = float(dawk.get())
    except ValueError:
        messagebox.showinfo("uwaga", "Błędnie podany parametr - niepoprawny znak")
    else:
        if czas.get() == ' sekunda':
            parametry[3] = 3600
        elif czas.get() == ' minuta':
            parametry[3] = 60
        elif czas.get() == ' godzina':
            parametry[3] = 1.0
        elif czas.get() == ' doba':
            parametry[3] = 1 / 24

        fun_z[5] = 0
        fun_d[7] = 0  # NIE UŻYWAĆ EFEKTU SĄSIEDZTWA - BŁĘDNA I KOSZMARNA IMPLEMENTACJA

        parametry[4] += 1
        if parametry[1] * parametry[1] * parametry[1] * parametry[2] < 120000000:
            if wczytywanie == 0:
                sym = Symulacja(zdrowa, uszkodzona, zmutowana, nowotworowa, wspolne, adapt,
                                sasiedztwo, fun_z, fun_u, fun_m, fun_n, parametry, dawka)
                sym.empty()
                sym.ustaw_status("zdrowa")
            else:
                sym = Symulacja(zdrowa, uszkodzona, zmutowana, nowotworowa, wspolne, adapt,
                                sasiedztwo, fun_z, fun_u, fun_m, fun_n, parametry, dawka)
                sym.empty()
                sym.ustaw_wczytane(organizm)
            zapis, org, zd, uszk, zm, nw, ws, wykr = sym.symulacja()
            wyswietl_wykres()

        else:
            # bo mi lapek wywala brak pamięci
            messagebox.showinfo("uwaga", "liczba kroków x (liczba komórek + 2)^3 nie może wynosić więcej niż 1,2*10^8")
            kroki.delete(0, END)
            kroki.insert(0, 13000)
            wszystkie_wart.delete(0, END)
            wszystkie_wart.insert(0, 20)
            if parametry[1] > 20:
                pocz_wart.delete(0, END)
                pocz_wart.insert(0, 20)


window = Tk()  # okno programu
window.geometry("550x420")
window.title("Symulacja")

menu_bar = Menu(window)
window.config(menu=menu_bar)
file_menu = Menu(menu_bar)
menu_bar.add_command(label="Wybór parametrów", command=lambda: set_parameters(window))
menu_bar.add_command(label="Wczytaj komórki", command=lambda: wczytaj_kom())
menu_bar.add_command(label="Wczytaj dawkę", command=lambda: wczytaj_daw())
menu_bar.add_command(label="Pomoc", command=lambda: pomoc(window))

middleFrame = Frame(window)
middleFrame.pack(side=TOP, fill=Y)
middleFrame.config(padx=10, pady=10)

used_parameters = LabelFrame(middleFrame, text="Parametry początkowe", width=300, height=170)  # ramka

pocz_wart = Entry(used_parameters)
pocz_wart.grid(row=0, column=0, padx=10, pady=4)
pocz_wart.config(width=10)
pocz_wart.insert(0, parametry[0])

label_pocz = Label(used_parameters, text='^3 - początkowa liczba komórek')
label_pocz.grid(row=0, column=1, padx=5, pady=4, sticky='w')

wszystkie_wart = Entry(used_parameters)
wszystkie_wart.grid(row=1, column=0, padx=10, pady=4)
wszystkie_wart.config(width=10)
wszystkie_wart.insert(0, parametry[1])

label_wszystkie = Label(used_parameters, text='^3 - liczba komórek')
label_wszystkie.grid(row=1, column=1, padx=5, pady=4, sticky='w')

kroki = Entry(used_parameters)
kroki.grid(row=2, column=0, padx=10, pady=4)
kroki.config(width=10)
kroki.insert(0, parametry[2])

label_kroki = Label(used_parameters, text='liczba kroków')
label_kroki.grid(row=2, column=1, padx=5, pady=4, sticky='w')

par = Entry(used_parameters)
par.grid(row=3, column=0, padx=10, pady=4)
par.config(width=10)
par.insert(0, parametry[5])

label_par = Label(used_parameters, text='promień')
label_par.grid(row=3, column=1, padx=5, pady=4, sticky='w')

dawk = Entry(used_parameters)
dawk.grid(row=4, column=0, padx=10, pady=4)
dawk.config(width=10)
dawk.insert(0, 0)

label_dawk = Label(used_parameters, text='dawka [Gy]')
label_dawk.grid(row=4, column=1, padx=5, pady=4, sticky='w')

used_parameters.grid(row=0, column=0, rowspan=5, columnspan=2, padx=5, pady=0, sticky='n')
used_parameters.grid_propagate(False)

# inne parametry
listBox = Listbox(master=middleFrame, height=2, selectmode=SINGLE, exportselection=False, width=23)
listBox.insert(1, "Limfocyty")
listBox.insert(2, "Wczytane")
listBox.grid(row=0, column=2, rowspan=3, padx=10, pady=20)
listBox.select_set(0)

n = StringVar()
czas = ttk.Combobox(middleFrame, textvariable=n, state="readonly")
czas['values'] = (' sekunda', ' minuta', ' godzina', ' doba')
czas.grid(column=2, row=4, padx=20, pady=10)
czas.current(2)

chart_data = LabelFrame(middleFrame, text="Dane do wykresu/zapisu", width=300, height=180)  # ramka

var_healthy = IntVar()
check_healthy = Checkbutton(chart_data, text="Zdrowe", variable=var_healthy)
check_healthy.select()
check_healthy.grid(row=0, column=0, padx=10, pady=1, sticky='w')

var_damaged = IntVar()
check_damaged = Checkbutton(chart_data, text="Uszkodzone", variable=var_damaged)
check_damaged.select()
check_damaged.grid(row=1, column=0, padx=10, pady=1, sticky='w')

var_mutated = IntVar()
check_mutated = Checkbutton(chart_data, text="Zmutowane", variable=var_mutated)
check_mutated.select()
check_mutated.grid(row=2, column=0, padx=10, pady=1, sticky='w')

var_cancerous = IntVar()
check_cancerous = Checkbutton(chart_data, text="Nowotworowe", variable=var_cancerous)
check_cancerous.select()
check_cancerous.grid(row=3, column=0, padx=10, pady=1, sticky='w')

var_dead = IntVar()
check_dead = Checkbutton(chart_data, text="Martwe", variable=var_dead)
check_dead.select()
check_dead.grid(row=4, column=0, padx=10, pady=1, sticky='w')

var_all = IntVar()
check_all = Checkbutton(chart_data, text="Żywe", variable=var_all)
check_all.grid(row=5, column=0, padx=10, pady=1, sticky='w')

chart_data.grid(row=6, column=0, rowspan=6, padx=2, pady=0, sticky='n')
chart_data.grid_propagate(False)

save = LabelFrame(middleFrame, text="Wyniki symulacji", width=150, height=180)  # ramka

zapWyn = Button(save, text="Zapisz wyniki", command=lambda: zapisz_dane())
zapWyn.grid(row=0, column=0, padx=10, pady=12)
zapWyn.config(width=15)

zapKom = Button(save, text="Zapisz komórki", command=lambda: zapisz_komorki())
zapKom.grid(row=1, column=0, padx=10, pady=12)
zapKom.config(width=15)

wysWyk = Button(save, text="Wyświetl wykres", command=lambda: wyswietl_wykres())
wysWyk.grid(row=3, column=0, padx=10, pady=12)
wysWyk.config(width=15)

save.grid(row=5, column=2, rowspan=2, padx=5, pady=0, sticky='n')
save.grid_propagate(False)

bottomFrame = Frame(window)
bottomFrame.pack(side=BOTTOM, fill=Y)
bottomFrame.config(padx=10, pady=0)

butRes = Button(bottomFrame, text="Reset", command=lambda: reset2())
butRes.grid(row=0, column=2, padx=10, pady=10)
butRes.config(width=10)

butNext = Button(bottomFrame, text="Dalej", command=lambda: dalej())
butNext.grid(row=0, column=1, padx=10, pady=10)
butNext.config(width=10)

window.mainloop()

